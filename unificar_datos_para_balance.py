#!/usr/bin/python3

import pandas as pd
import openpyxl
import os
# from openpyxl import workbook, load_workbook

mesanterior = int(pd.to_datetime("today").strftime("%Y%m"))-1
# mesanterior = int(pd.to_datetime("today").strftime("%Y%m"))
# mesanterior = '202108'

os.chdir('/media/trabajo/Trabajo/WEME/exportacion/estudio')
dir = '/media/trabajo/Trabajo/WEME/exportacion/estudio/'

ventas = f'{dir}{mesanterior}ivaventas.xls'
compras = f'{dir}{mesanterior}ivacpra.xls'
deudores = f'{dir}{mesanterior}deudores.xls'
proveedores = f'{dir}{mesanterior}prov.xls'
prod = f'{dir}{mesanterior}prod.xls'

for planilla in ventas, compras, prod, deudores, proveedores:
    # os.system(f'libreoffice --convert-to xlsx {planilla}')
    comando ='libreoffice7.2 --convert-to xlsx ' + planilla
    os.system(comando)



# ivaventas = openpyxl.load_workbook(ventas+'x')
# ivacompras = openpyxl.load_workbook(compras+'x')
# produccion = openpyxl.load_workbook(prod+'x')
# deudoresxventas = openpyxl.load_workbook(deudores+'x')
# deudaproveedores = openpyxl.load_workbook(proveedores+'x')

ivaventas = pd.read_excel(ventas+'x',engine='openpyxl')
ivacompras = pd.read_excel(compras+'x',engine='openpyxl')
produccion = pd.read_excel(prod+'x',engine='openpyxl')
deudoresxventas = pd.read_excel(deudores+'x',engine='openpyxl')
deudaproveedores = pd.read_excel(proveedores+'x',engine='openpyxl')

# hojaventas = ivaventas.sheetnames[0]
# hojaventas.active
# print(hojaventas)
# min_columna = ivaventas.active.min_column
# min_fila = ivaventas.active.min_row
# max_columna = ivaventas.active.max_column
# max_fila = ivaventas.active.max_row

#  SACO LAS COLUMNAS QUE QUIERO DE LOS DATAFRAMES
ivaventas = ivaventas[['fecha', 'documento', 'razon_social', 'neto']]
ivacompras = ivacompras[['fecha', 'documento', 'razon_social', 'neto']]
produccion = produccion[['codigo','descripcion','venta','ppromedio']]
deudoresxventas = deudoresxventas[['cliente', 'emision', 'documento', 'saldo']]
deudaproveedores = deudaproveedores[['proveedor', 'emision', 'documento', 'saldo']]

#  LIMPIO LOS DATAFRAMES
# df = df.drop(df[df['C']==True].index) 
ivaventas = ivaventas.drop(ivaventas[ivaventas['razon_social']=='LATIN CHEMICAL SUPPLIERS S.A.'].index) 
ivaventas = ivaventas.drop(ivaventas[ivaventas['fecha']=='  -   -'].index) 
ivacompras = ivacompras.drop(ivacompras[ivacompras['razon_social']=='MOLINO CAMPODONICO'].index) 
ivacompras = ivacompras.drop(ivacompras[ivacompras['razon_social']=='COOP DE TRAB MOLIN  SALAD LTDA'].index) 
ivacompras = ivacompras.drop(ivacompras[ivacompras['razon_social']=='CA-MI SRL'].index) 
ivacompras = ivacompras.drop(ivacompras[ivacompras['razon_social']=='EMPRESA DISTRIBUIDORA SUR SA'].index) 
deudoresxventas = deudoresxventas.drop(deudoresxventas[deudoresxventas['cliente']=='Total'].index) 
deudoresxventas = deudoresxventas.drop(deudoresxventas[deudoresxventas['cliente']=='TEMPO ABASTO'].index) 
deudoresxventas = deudoresxventas.drop(deudoresxventas[deudoresxventas['cliente']=='ARMENIA 1231'].index) 
deudoresxventas = deudoresxventas.drop(deudoresxventas[deudoresxventas['cliente']=='MARSANO HNOS SA'].index) 
deudoresxventas = deudoresxventas.drop(deudoresxventas[deudoresxventas['cliente']=='LATIN CHEMICAL SUPPLIERS S.A.'].index) 
deudoresxventas = deudoresxventas.drop(deudoresxventas[deudoresxventas['cliente']=='INTERCARGO'].index) 
deudoresxventas = deudoresxventas.drop(deudoresxventas[deudoresxventas['cliente']=='KIMBERLEY PILAR'].index) 
deudoresxventas = deudoresxventas.drop(deudoresxventas[deudoresxventas['cliente']=='IMPRESORES PILAR'].index) 
deudoresxventas = deudoresxventas.drop(deudoresxventas[deudoresxventas['cliente']=='AGROALIMENTOS LA PALMA SRL'].index) 
deudoresxventas = deudoresxventas.drop(deudoresxventas[deudoresxventas['cliente']=='SINDICATO PANADEROS'].index) 
deudoresxventas = deudoresxventas.drop(deudoresxventas[deudoresxventas['cliente']=='CARREFOUR CUENTA MADRE'].index) 
deudoresxventas = deudoresxventas.drop(deudoresxventas[deudoresxventas['cliente']=='CENTRO DISTRIBUCION'].index) 
deudoresxventas = deudoresxventas.drop(deudoresxventas[deudoresxventas['emision']=='  -   -'].index) 
deudaproveedores = deudaproveedores.drop(deudaproveedores[deudaproveedores['proveedor']=='Total'].index) 
deudaproveedores = deudaproveedores.drop(deudaproveedores[deudaproveedores['proveedor']=='DEVISUR SA'].index) 
deudaproveedores = deudaproveedores.drop(deudaproveedores[deudaproveedores['proveedor']=='GIDEA SRL'].index) 
deudaproveedores = deudaproveedores.drop(deudaproveedores[deudaproveedores['proveedor']=='PAMBUKIAN E IZZO SRL'].index) 
deudaproveedores = deudaproveedores.drop(deudaproveedores[deudaproveedores['proveedor']=='POLIZYM S.A.C.I.'].index) 
deudaproveedores = deudaproveedores.drop(deudaproveedores[deudaproveedores['proveedor']=='MOLINO CAMPODONICO'].index) 
deudaproveedores = deudaproveedores.drop(deudaproveedores[deudaproveedores['proveedor']=='SATVA SRL'].index) 
deudaproveedores = deudaproveedores.drop(deudaproveedores[deudaproveedores['proveedor']=='SATVA SRL'].index) 
deudaproveedores = deudaproveedores.drop(deudaproveedores[deudaproveedores['emision']=='  -   -'].index) 
produccion = produccion.drop(produccion[produccion['descripcion']=='VIANDAS'].index)

#  TOTALES
# total_general.loc['Total General'] = impagas_recibos_nc[['Total Factura', 'Saldo Factura']].sum()
# deudaproveedores.loc['Total General'] = deudaproveedores[['saldo']].sum()
# deudoresxventas.loc['Total General'] = deudoresxventas[['saldo']].sum()
# ivacompras.loc['Total General'] = ivacompras[['neto']].sum()
# ivaventas.loc['Total General'] = ivaventas[['neto']].sum()


# CREAMOS EL ESCRITOR
writer = pd.ExcelWriter(f'/media/trabajo/Trabajo/Administracion/Varios/DatosBalance/{mesanterior}DatosBalance.xlsx')
# ESCRIBIMOS EL DATASFREME A LA HOJA CON X NOMBRE
ivaventas.to_excel(writer, 'IVAVENTAS', index=False)
ivacompras.to_excel(writer, 'IVACOMPRAS', index=False)
produccion.to_excel(writer, 'PRODUCCION', index=False)
deudoresxventas.to_excel(writer, 'DEUDORES', index=False)
deudaproveedores.to_excel(writer, 'PROVEEDORES', index=False)
# save the excel file
writer.save()

for planilla in ventas, compras, prod, deudores, proveedores:
    os.system(f"rm {planilla}"+'x')

# print(ivaventas)
# print(ivacompras)
# print(produccion)
# print(deudoresxventas)
# print(deudaproveedores)
