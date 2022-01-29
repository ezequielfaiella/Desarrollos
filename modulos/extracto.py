#!/usr/bin/python3
#  NECESITA            xlrd  1.2.0
import pandas as pd
import openpyxl
import os
from openpyxl import workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import datetime
import logging
from time import sleep

logger = logging
logger.basicConfig(level=logging.INFO,
                   format='%(asctime)s: %(levelname)s [%(filename)s:%(lineno)s] %(message)s',
                   datefmt='%Y%m%d-%I%M%S-%p',
                   handlers=[
                    #    log en la terminal
                       logging.StreamHandler(),
                    #    log en archivo
                       logging.FileHandler(__file__+'.log')
                   ])

# from modulos import extracto

# UBICACION ARCHIVO DE LIBRO BANCO
# os.chdir('/home/ezequiel/Escritorio/')
os.chdir('/media/trabajo/Trabajo/Administracion/1 Uso Diario')
#  NOMBRE DE ARCHIVO DE LIBRO BANCO
banco = 'BancoFrances.xlsx'

fecha = datetime.datetime.today().strftime("%d-%m-%Y")

# print(fecha)

#  FUNCION PARA ANEXAR EL DATAFRAME
def paso_a_excel(dataframe, hoja, fila, columna):
    ''' 
        uso de la funcion que incorpora un dataframe de pandas a un libro de excel (en revision ods) ya existente.
        uso: paso_a_excel(dataframe al que se quiere pasar, nombre de la hoja, fila de inicio del pagado, columna
        de inicio del pegado de la info)
    '''
    libro = load_workbook(banco)
    hoja = libro[hoja]
    # hoja = libro.get_sheet_by_name(hoja)

    rows = dataframe_to_rows(dataframe, index=False, header=True)

    for r_idx, row in enumerate(rows, int(fila)):  # 1 es la fila donde quiero que comienze
        for c_idx, value in enumerate(row, int(columna)):   # 1 es la columna x la qu quiero que empiece
            hoja.cell(row=r_idx, column=c_idx, value=value)

    libro.save(banco)

def fun_emitidos(emitidos,libro_banco_emitidos):
        # #  ACTUALIZACION DE LOS CHEQUES EMITIDOS
        logger.info('inicia proceso de comparacion de echeq emitidos')
        if emitidos.shape[0] > 1:
            try:
                emitidos['Fecha de emisión']=pd.to_datetime(emitidos['Fecha de emisión'], format="%d/%m/%Y")
                libro_banco_emitidos['Fecha de emisión']=pd.to_datetime(libro_banco_emitidos['Fecha de emisión'], format="%d/%m/%Y")
                emitidos['Fecha de Pago']=pd.to_datetime(emitidos['Fecha de Pago'], format="%d/%m/%Y")
                emitidos['Fecha de vencimiento']=pd.to_datetime(emitidos['Fecha de vencimiento'], format="%d/%m/%Y")
                libro_banco_emitidos['Fecha de vencimiento']=pd.to_datetime(libro_banco_emitidos['Fecha de vencimiento'], format="%d/%m/%Y")
                emitidos['Monto'] = pd.to_numeric(emitidos['Monto'].str.replace('$', '', regex=False))
                libro_banco_emitidos['N° de cheque']=libro_banco_emitidos['N° de cheque'].astype('Int64',errors='ignore')
                emitidos['N° de cheque']=emitidos['N° de cheque'].astype('Int64',errors='ignore')
                libro_banco_emitidos['N° de cuenta de emisión']=libro_banco_emitidos['N° de cuenta de emisión'].astype('Int64',errors='ignore')
                emitidos['N° de cuenta de emisión']=emitidos['N° de cuenta de emisión'].astype('Int64',errors='ignore')

                #  mezclo los dos dataframes
                emitidos_faltantes = libro_banco_emitidos.merge(emitidos, how='outer')

                if emitidos.empty:
                    print('Sin Emitidos para agregar')
                    pass
                else:
                    emitidos_faltantes['Fecha de emisión']=pd.to_datetime(emitidos_faltantes['Fecha de emisión'], format="%d/%m/%Y")
                    # aceptados_faltantes['Fecha de depósito']=pd.to_datetime(aceptados_faltantes['Fecha de depósito'], format="%d/%m/%Y")
                    emitidos_faltantes['Fecha de Pago']=pd.to_datetime(emitidos_faltantes['Fecha de Pago'], format="%d/%m/%Y")
                    emitidos_faltantes['Fecha de vencimiento']=pd.to_datetime(emitidos_faltantes['Fecha de vencimiento'], format="%d/%m/%Y")
                    # depositos_faltantes['Monto'] = pd.to_numeric(depositos_faltantes['Monto'].str.replace('$', '', regex=False))
                    emitidos_faltantes = emitidos_faltantes.sort_values('N° de cheque')
                    emitidos_faltantes = emitidos_faltantes[emitidos_faltantes['Emitido a'].notna()] # BORRA LA FILA SI NO TIENE "EMITIDO A"
                    # print(emitidos_faltantes)
                    paso_a_excel(emitidos_faltantes,'Echeq', 2, 1)  
            except Exception as e: 
                logger.error(f'fallo en proceso fun_emitidos. {e}')
                print(e, '\nFallo al actualizar los echeq emitidos\n')
                # emitidos.info()
                # libro_banco_emitidos.info()
                pass

def fun_aceptados(aceptados,libro_banco_aceptados):
        #  ACTUALIZACION DE LOS CHEQUES EN CARTERA
        # if aceptados:
            # if aceptados.shape[0] > 1:
            logger.info('inicia proceso de comparacion de echeq aceptados')
            if not aceptados.empty:
                try:
                    # libro_banco_aceptados = libro_banco_aceptados.drop_duplicates(subset=None, keep='first', inplace=False, ignore_index=False)
                    # print(libro_banco_aceptados)
                    aceptados['Fecha de emisión']=pd.to_datetime(aceptados['Fecha de emisión'], format="%d/%m/%Y")
                    libro_banco_aceptados['Fecha de emisión']=pd.to_datetime(libro_banco_aceptados['Fecha de emisión'], format="%d/%m/%Y")
                    aceptados['Fecha de Pago']=pd.to_datetime(aceptados['Fecha de Pago'], format="%d/%m/%Y")
                    libro_banco_aceptados['Fecha de Pago']=pd.to_datetime(libro_banco_aceptados['Fecha de Pago'], format="%d/%m/%Y")
                    aceptados['Fecha de vencimiento']=pd.to_datetime(aceptados['Fecha de vencimiento'], format="%d/%m/%Y")
                    libro_banco_aceptados['Fecha de vencimiento']=pd.to_datetime(libro_banco_aceptados['Fecha de vencimiento'], format="%d/%m/%Y")
                    aceptados['Monto'] = pd.to_numeric(aceptados['Monto'].str.replace('$', '', regex=False))
                    # libro_banco_aceptados['Monto'] = pd.to_numeric(libro_banco_aceptados['Monto'].str.replace('$', '', regex=False))
                    aceptados['Tipo de Documento del Emisor']=aceptados['Tipo de Documento del Emisor'].astype(str)
                    libro_banco_aceptados['Tipo de Documento del Emisor']=libro_banco_aceptados['Tipo de Documento del Emisor'].astype(str)
                    aceptados['Endosos']=aceptados['Endosos'].astype(str)
                    libro_banco_aceptados['Endosos']=libro_banco_aceptados['Endosos'].astype(str)


                    #  mezclo los dos dataframes

                    aceptados_faltantes = libro_banco_aceptados.merge(aceptados, how='outer')

                    aceptados_faltantes['Fecha de emisión']=pd.to_datetime(aceptados_faltantes['Fecha de emisión'], format="%d/%m/%Y")
                    # aceptados_faltantes['Fecha de depósito']=pd.to_datetime(aceptados_faltantes['Fecha de depósito'], format="%d/%m/%Y")
                    aceptados_faltantes['Fecha de Pago']=pd.to_datetime(aceptados_faltantes['Fecha de Pago'], format="%d/%m/%Y")
                    aceptados_faltantes['Fecha de vencimiento']=pd.to_datetime(aceptados_faltantes['Fecha de vencimiento'], format="%d/%m/%Y")
                    # depositos_faltantes['Monto'] = pd.to_numeric(depositos_faltantes['Monto'].str.replace('$', '', regex=False))
                    aceptados_faltantes = aceptados_faltantes.sort_values('Fecha de Pago')
                    aceptados_faltantes = aceptados_faltantes[aceptados_faltantes['Emitido por'].notna()] # BORRA LA FILA SI NO TIENE "EMITIDO A"
                    # print(aceptados_faltantes)
                    # aceptados_faltantes.to_excel('con_duplicados.xlsx')
                    aceptados_faltantes = aceptados_faltantes.drop_duplicates(subset='CheqID', keep='first', inplace=False, ignore_index=False)
                    # aceptados_faltantes.to_excel('sin_duplicados.xlsx')
                    # print(aceptados_faltantes)
                    # exit()
                    paso_a_excel(aceptados_faltantes,'Cartera', 1, 1)  
                    # aceptados_faltantes.to_excel('nuevo.ods')  
                except Exception as e: 
                    logger.error(f'fallo en proceso fun_aceptados. {e}')
                    print(e, '\nFallo al actualizar los cheques en cartera\n')
                    aceptados.info()
                    libro_banco_aceptados.info()
                    # print(aceptados)
                    # print(aceptados_faltantes)
                    # print(libro_banco_aceptados)
                    pass

def fun_depositos(deposito,libro_banco_depositados):
        #  ACTUALIZACION DE LOS DEPOSITOS
        logger.info('inicia proceso de comparacion de depositos')
        if deposito.shape[0] > 1:
            try:
                logger.info('empiezo a ajustar los datos')
                # acomodo los datos para poder mezclar
                deposito['Fecha de emisión']=pd.to_datetime(deposito['Fecha de emisión'], format="%d/%m/%Y")
                deposito['Fecha de depósito']=pd.to_datetime(deposito['Fecha de depósito'], format="%d/%m/%Y")
                deposito['Fecha de pago']=pd.to_datetime(deposito['Fecha de pago'], format="%d/%m/%Y")
                deposito['Fecha de vencimiento']=pd.to_datetime(deposito['Fecha de vencimiento'], format="%d/%m/%Y")
                libro_banco_depositados['Fecha de emisión']=pd.to_datetime(libro_banco_depositados['Fecha de emisión'], format="%d/%m/%Y")
                libro_banco_depositados['Fecha de depósito']=pd.to_datetime(libro_banco_depositados['Fecha de depósito'], format="%d/%m/%Y")
                libro_banco_depositados['Fecha de pago']=pd.to_datetime(libro_banco_depositados['Fecha de pago'], format="%d/%m/%Y")
                libro_banco_depositados['Fecha de vencimiento']=pd.to_datetime(libro_banco_depositados['Fecha de vencimiento'], format="%d/%m/%Y")
                deposito['Monto'] = pd.to_numeric(deposito['Monto'].str.replace('$', '', regex=False))
                # libro_banco_depositados['Monto'] = pd.to_numeric(libro_banco_depositados['Monto'].str.replace('$', '', regex=False))

                #  mezclo los dos dataframes
                logger.info('hago el merge de los depósitos')
                depositos_faltantes = libro_banco_depositados.merge(deposito, how='outer')
                # reacomodo los tipos de dato
                logger.info('ajusto nuevamente los datos')
                depositos_faltantes['Fecha de emisión']=pd.to_datetime(depositos_faltantes['Fecha de emisión'], format="%d/%m/%Y")
                depositos_faltantes['Fecha de depósito']=pd.to_datetime(depositos_faltantes['Fecha de depósito'], format="%d/%m/%Y")
                depositos_faltantes['Fecha de pago']=pd.to_datetime(depositos_faltantes['Fecha de pago'], format="%d/%m/%Y")
                depositos_faltantes['Fecha de vencimiento']=pd.to_datetime(depositos_faltantes['Fecha de vencimiento'], format="%d/%m/%Y")
                # depositos_faltantes['Monto'] = pd.to_numeric(depositos_faltantes['Monto'].str.replace('$', '', regex=False))
                
                # ordeno los datos
                logger.info('ordeno por fecha de emision')
                depositos_faltantes = depositos_faltantes.sort_values('Fecha de emisión')

                # paso el dataframe final a la planilla
                logger.info('lo mando al excel')
                paso_a_excel(depositos_faltantes,'Depositados', 1, 1)    
            except Exception as e: 
                logger.error(f'fallo en proceso fun_depositos. {e}')
                print(e, '\nFallo al actualizar los depositos\n')
                # print('\ndataframe deposito\n')
                # deposito.info()
                # print('\ndataframe libro deposito\n')
                # libro_banco_depositados.info()
                pass

def fun_extracto_original_no_uso_mas(libro_banco_extracto,extracto_file): # ver si anda sin extracto que lo saque. con funciona
    logger.info('inicia proceso de comparacion de extracto')
    if len(pd.read_excel(extracto_file, sheet_name=None)) > 1:
        extracto1 = pd.read_excel(extracto_file, sheet_name='Movimientos Históricos', usecols='A:I', skiprows=6)
        extracto2 = pd.read_excel(extracto_file, sheet_name='Movimientos del Día', usecols='A:I', skiprows=6)
        extracto2 = extracto2.rename(columns={'Nro de cheque': 'Número Documento', 'Saldo Parcial': 'Detalle'})
        extracto1['Número Documento']=extracto1['Número Documento'].astype('int64',errors='ignore')
        extracto2['Detalle']=extracto2['Detalle'].astype(str,errors='ignore')
        extracto = pd.concat([extracto1, extracto2])
    else:
        extracto2 = pd.read_excel(extracto_file, sheet_name='Movimientos Históricos', usecols='A:I', skiprows=6)
        extracto2 = extracto2.rename(columns={'Nro de cheque': 'Número Documento', 'Saldo Parcial': 'Detalle'})
        extracto2['Detalle']=extracto2['Detalle'].astype(str,errors='ignore')
        extracto = extracto2
    # continua proceso unificado
    extracto['Codigo']=extracto['Codigo'].astype('float64',errors='ignore')
    extracto['Fecha']=pd.to_datetime(extracto['Fecha'], format="%d-%m-%Y", errors='ignore')
    extracto['Fecha Valor']=pd.to_datetime(extracto['Fecha Valor'], format="%d-%m-%Y", errors='ignore')
    # extracto['Fecha']=extracto['Fecha'].astype('datetime64[s]')
    # extracto['Fecha Valor']=extracto['Fecha Valor'].astype('datetime64[s]')
    extracto = extracto.sort_values('Fecha')
    extracto[['Detalle']]=extracto[['Detalle']].fillna('', inplace=True)
    #  con merge funciona bien pero tenia problemas q me duplicaba filas no se por que
    # saco solo los que estan en el extracto recien descargado
    extracto_faltantes = libro_banco_extracto.merge(extracto, how='outer', indicator=True).loc[lambda x: x['_merge'] == 'right_only']
    # print(extracto_faltantes)
    # extracto_faltantes.to_excel('comparacion.xlsx')
    # exit()
    extracto_faltantes = extracto_faltantes.drop(columns=['_merge'])
    #  uno el libro con el extracto
    extracto_faltantes = pd.concat([libro_banco_extracto, extracto_faltantes])
    #  borro lo vacio
    extracto_faltantes = extracto_faltantes[extracto_faltantes['Concepto'].notna()] # BORRA LA FILA SI NO TIENE "Concepto"
    #  lo paso al excel del banco
    paso_a_excel(extracto_faltantes, 'BANCO', 5, 1)        

def fun_extracto(libro_banco_extracto,extracto_file): # ver si anda sin extracto que lo saque. con funciona
    logger.info('inicia proceso de comparacion de extracto')
    if len(pd.read_excel(extracto_file, sheet_name=None)) > 1:
        extracto1 = pd.read_excel(extracto_file, sheet_name='Movimientos Históricos', usecols='A:I', skiprows=6)
        extracto2 = pd.read_excel(extracto_file, sheet_name='Movimientos del Día', usecols='A:I', skiprows=6)
        extracto2 = extracto2.rename(columns={'Nro de cheque': 'Número Documento', 'Saldo Parcial': 'Detalle'})
        extracto1['Número Documento']=extracto1['Número Documento'].astype('int64',errors='ignore')
        extracto2['Detalle']=extracto2['Detalle'].astype(str,errors='ignore')
        extracto = pd.concat([extracto1, extracto2])
    else:
        extracto2 = pd.read_excel(extracto_file, sheet_name='Movimientos Históricos', usecols='A:I', skiprows=6)
        extracto2 = extracto2.rename(columns={'Nro de cheque': 'Número Documento', 'Saldo Parcial': 'Detalle'})
        extracto2['Detalle']=extracto2['Detalle'].astype(str,errors='ignore')
        extracto = extracto2
    # continua proceso unificado
    extracto['Codigo']=extracto['Codigo'].astype('float64',errors='ignore')
    extracto['Fecha']=pd.to_datetime(extracto['Fecha'], format="%d-%m-%Y", errors='ignore')
    extracto['Fecha Valor']=pd.to_datetime(extracto['Fecha Valor'], format="%d-%m-%Y", errors='ignore')
    extracto = extracto.sort_values('Fecha')
    extracto[['Detalle']]=extracto[['Detalle']].fillna('', inplace=True)
    ultima_fecha_cargada = libro_banco_extracto['Fecha'].values[-1]
    extracto_final = extracto[extracto.Fecha >= ultima_fecha_cargada]
    libro_banco_extracto = libro_banco_extracto[libro_banco_extracto.Fecha < ultima_fecha_cargada]
    extracto_faltantes = pd.concat([libro_banco_extracto, extracto_final])
    paso_a_excel(extracto_faltantes, 'BANCO', 5, 1)   

def fun_endosados(endosados,libro_banco_endosados):
        #  ACTUALIZACION DE LOS DEPOSITOS
        logger.info('inicia proceso de comparacion de endosos')
        if endosados.shape[0] > 1:
            try:
                #  mezclo los dos dataframes
                logger.info('hago el merge de los depósitos')
                endosos_faltantes = libro_banco_endosados.merge(endosados, how='outer')
                # paso el dataframe final a la planilla
                logger.info('lo mando al excel')
                paso_a_excel(endosos_faltantes,'Endosados', 1, 1)    
            except Exception as e: 
                logger.error(f'fallo en proceso fun_endosos. {e}')
                print(e, '\nFallo al actualizar los endosos\n')
                pass

def fun_cedidos(cedidos,libro_banco_cedidos):
        #  ACTUALIZACION DE LOS DEPOSITOS
        logger.info('inicia proceso de comparacion de cedidos')
        if cedidos.shape[0] > 1:
            try:
                #  mezclo los dos dataframes
                logger.info('hago el merge de los depósitos')
                cedidos_faltantes = libro_banco_cedidos.merge(cedidos, how='outer')
                # paso el dataframe final a la planilla
                logger.info('lo mando al excel')
                paso_a_excel(cedidos_faltantes,'Cedidos', 1, 1)    
            except Exception as e: 
                logger.error(f'fallo en proceso fun_cedisos. {e}')
                print(e, '\nFallo al actualizar los cedidos\n')
                pass
      
# if __name__ == '__main__':
def proceso():
    
    logger.info('inicia proceso de chequeo de existencia de archivos')
    
    if os.path.exists(f'/tmp/{fecha}-eCheqs-deposito.xls'):
        deposito_file = f'/tmp/{fecha}-eCheqs-deposito.xls'       # va en depositados del libro banco
    else:
        deposito_file = ''
        logger.info(f'no existe el archivo {fecha}-eCheqs-deposito.xls')
        
    if os.path.exists(f'/tmp/{fecha}-eCheqs-aceptados.xls'):
        aceptados_file = f'/tmp/{fecha}-eCheqs-aceptados.xls'     # va en cartera del libro banco        
    else:
        aceptados_file = ''
        logger.info(f'no existe el archivo {fecha}-eCheqs-aceptados.xls')
        
    if os.path.exists(f'/tmp/{fecha}-eCheqs-emitidos.xls'):
        emitidos_file = f'/tmp/{fecha}-eCheqs-emitidos.xls'       # va en echeq del libro banco
    else:
        emitidos_file = ''
        logger.info(f'no existe el archivo {fecha}-eCheqs-emitidos.xls')
    
    if os.path.exists('/tmp/Movimientos.xls'):
        extracto_file = '/tmp/Movimientos.xls'
    else:
        extracto_file = ''
        logger.info(f'no existe el archivo Movimientos.xls')
    
    
    if os.path.exists(f'/tmp/{fecha}-eCheqs-cedidos.xls'):
        cedidos_file = f'/tmp/{fecha}-eCheqs-cedidos.xls'
    else:
        cedidos_file = ''
        logger.info(f'no existe el archivo {fecha}-eCheqs-cedidos.xls')

    if os.path.exists(f'/tmp/{fecha}-eCheqs-endosados.xls'):
        endosados_file = f'/tmp/{fecha}-eCheqs-endosados.xls'
    else:
        endosados_file = ''
        logger.info(f'no existe el archivo {fecha}-eCheqs-endosados.xls')
                
    #  convierto a xlsx por si no no me las quiere abrir. por ahora no se usa
    # for planilla in aceptados, deposito, emitidos:
    #     os.system(f'libreoffice7.1 --convert-to xlsx {planilla}')

    #  cargo con pandas los dataframes originales
    logger.info(f'abre libro banco en {banco}')
    # libro_banco = pd.ExcelFile(banco, engine='openpyxl')

    #  como sacar un rango de una hoja
    # pd.read_excel('resultat-elections-2012.xls', index_col = None, skiprows= 2, nrows= 5, sheet_name='France entière T1T2', usecols=range(0,8))

    logger.info('inicia proceso de conversion de datos para unificar tipos')
    
    if deposito_file:
        # libro_banco_depositados = libro_banco.parse(sheet_name='Depositados', header=0, usecols='A:Q')
        libro_banco_depositados = pd.read_excel(banco,sheet_name='Depositados', header=0, usecols='A:Q')
        deposito = pd.read_excel(deposito_file, usecols='A:Q')
        for df in deposito, libro_banco_depositados:
            df.convert_dtypes()
        logger.info('inicia fun_depositos')
        fun_depositos(deposito,libro_banco_depositados)
        sleep(3)
      
    if aceptados_file:
        # libro_banco_aceptados = libro_banco.parse(sheet_name='Cartera', header=0, usecols='A:N')
        libro_banco_aceptados = pd.read_excel(banco,sheet_name='Cartera', header=0, usecols='A:N')
        aceptados = pd.read_excel(aceptados_file, usecols='A:N')
        for df in aceptados, libro_banco_aceptados:
            df.convert_dtypes()
        logger.info('inicia fun_aceptados')
        fun_aceptados(aceptados,libro_banco_aceptados)
        sleep(3)
               
    if emitidos_file:
        # libro_banco_emitidos = libro_banco.parse(sheet_name='Echeq', header=1, usecols='A:I')
        libro_banco_emitidos = pd.read_excel(banco,sheet_name='Echeq', header=1, usecols='A:I')
        emitidos = pd.read_excel(emitidos_file, usecols='A:I')
        for df in emitidos, libro_banco_emitidos:
            df.convert_dtypes()
        logger.info('inicia fun_emitidos')
        fun_emitidos(emitidos,libro_banco_emitidos)
        sleep(3)
    
    if extracto_file:
        # libro_banco_extracto = libro_banco.parse(sheet_name='BANCO', header=4, usecols='A:I')
        libro_banco_extracto = pd.read_excel(banco,sheet_name='BANCO', header=4, usecols='A:I')
        extracto = pd.read_excel(extracto_file, usecols='A:I')
        for df in extracto, libro_banco_extracto:
            df.convert_dtypes()
        logger.info('inicia fun_extracto')
        fun_extracto(libro_banco_extracto,extracto_file) # ver si anda sin extracto que lo saque. con funciona
        sleep(3)

        
    if cedidos_file:
        # libro_banco_cedidos = libro_banco.parse(sheet_name='Cedidos', header=0, usecols='A:O')
        libro_banco_cedidos = pd.read_excel(banco,sheet_name='Cedidos', header=0, usecols='A:O')
        cedidos = pd.read_excel(cedidos_file, usecols='A:O')
        for df in cedidos, libro_banco_cedidos:
            df.convert_dtypes()
        logger.info('inicia fun_cedidos')
        fun_cedidos(cedidos,libro_banco_cedidos)
        sleep(3)
    
    if endosados_file:
        # libro_banco_endosados = libro_banco.parse(sheet_name='Endosados', header=0, usecols='A:P')
        libro_banco_endosados = pd.read_excel(banco,sheet_name='Endosados', header=0, usecols='A:P')
        endosados = pd.read_excel(endosados_file, usecols='A:P')
        for df in endosados, libro_banco_endosados:
            df.convert_dtypes()
        # libro_banco_endosados.info()
        # endosados.info()
        logger.info('inicia fun_endosados')
        fun_endosados(endosados,libro_banco_endosados) 
        sleep(3)
        
    # libro_banco.close()
    logger.info('carga en openpyxl el archivo de banco para darle formato')
    librobanco = openpyxl.load_workbook(banco)
    ### formatos
    formatofecha = NamedStyle(name='fecha', number_format='%d/%m/%Y')
    header = NamedStyle(name="header")
    header.font = Font(bold=True)
    header.border = Border(bottom=Side(border_style="thin"))
    header.alignment = Alignment(horizontal="center", vertical="center")

    ###
    hoja_banco = librobanco['BANCO']
    hoja_echeq = librobanco['Echeq']
    hoja_cartera = librobanco['Cartera']
    # primercelda_extracto = 3300
    
    # me hace ruido que me haya dado error y por eso puse el else. pero si no pienso mal siempre hay una ultima, con lo que no deberia dar error.
    # ver de poner un control que si ultima es igual a primera no haga nada porque deberia ser que no se actualizo nada.
    for cell in hoja_banco["M"]:
        if cell.row >6:
            if cell.value is None:
                # print(cell.row)
                primercelda_extracto = cell.row
                break
            else:
                primercelda_extracto = hoja_banco.max_row

    ultimacelda_extracto = hoja_banco.max_row
    # print(celdalibre)
    # print(banco['A1'].value)
    # for row in banco.iter_rows(min_col=1, max_col=2):
    #     for cell in row:
    #         cell.style = formatofecha

    for row in hoja_banco.iter_rows(min_col=1, max_col=2, min_row=6):
        for cell in row:
            # cell.style = formatofecha
            # cell.value = cell.value.strftime('%d/%m/%Y')
            cell.number_format = 'DD/MM/YYYY'

    ultimacelda_echeq = hoja_echeq.max_row
    for row in hoja_echeq.iter_rows(min_col=3, max_col=5, min_row=3, max_row=ultimacelda_echeq):
        for cell in row:
            # cell.style = formatofecha
            # cell.value = cell.value.strftime('%d/%m/%Y')
            cell.number_format = 'DD/MM/YYYY'
    
    ultimacelda_cartera = hoja_cartera.max_row
    for row in hoja_cartera.iter_rows(min_col=4, max_col=6, min_row=2, max_row=ultimacelda_cartera):
        for cell in row:
            # cell.style = formatofecha
            # cell.value = cell.value.strftime('%d/%m/%Y')
            cell.number_format = 'DD/MM/YYYY'

    # esta parte deberia de copiar la formula para saber el estado del cheque, no me la toma correctamente
    # for row in range(2, ultimacelda_cartera+1):
    #         # print(ultimacelda_cartera)
    #         # n = '=if(iseRROR(match(H%d;$Depositados!I:I;0));If(isERROR(match(H%d;$Cedidos!G:G;0));if(isERROR(match(H%d;$Endosados!H:H;0));"Cartera";"Endosado");"Cedido");"Depositado")' % (row, row, row)
    #         n = '=SI(ESERROR(COINCIDIR(H%d;Depositados!I:I;0));SI(ESERROR(COINCIDIR(H%d;Cedidos!G:G;0));SI(ESERROR(COINCIDIR(H%d;Endosados!H:H;0));"Cartera";"Endosado");"Cedido");"Depositado")' % (row, row, row)
    #         hoja_cartera['O' + str(row)].value = n
    #         # print(hoja_cartera['O' + str(row)].value)
            
    # for row, cell in enumerate(list(banco.columns)[10]):
    #     # print(cell)
    #     if row > primercelda and row < ultimacelda:
    #         n = '=(E%d-F%d)/C%d' % (row, row, row)
    #         print(cell, n) # check that n gets assigned correct string value
    #         cell.value = n
    #         # cellObj.value = 'inserta formula'
          
    #    SUBTOTAL
    for row in range(primercelda_extracto, ultimacelda_extracto+1):
        n = '=+K%d+G%d+H%d' % (row-1, row, row)
        hoja_banco['K' + str(row)].value = n

    #  SUBCATEGORIA
    for row in range(primercelda_extracto, ultimacelda_extracto+1):
    # # =SI.ND(BUSCARV(D3274;$Hoja2.$A$1:$C$304;3;0);"SIN CLASIFICAR")
    # # _xlnf.  se pone si no reconoce la funcion que le pasamos
    #     n = f'=_xlfn.IFNA(VLOOKUP(D{row};Hoja2!$A$1:$C$304;3;0);"SIN CLASIFICAR")'
        n = '=_xlfn.IFNA(VLOOKUP(D%d,Hoja2!$A$1:$C$304,3,0),"SIN CLASIFICAR")' % (row)
        hoja_banco['J' + str(row)].value = n
        # print(hoja_banco['J' + str(row)].value)

    #  CATEGORIA
    for row in range(primercelda_extracto, ultimacelda_extracto+1):
    # # =SI.ND(BUSCARV(D3274;$Hoja2.$A$1:$B$304;2;0);"SIN CLASIFICAR")
        n = '=_xlfn.IFNA(VLOOKUP(D%d,Hoja2!$A$1:$B$304,2,0),"SIN CLASIFICAR")' % (row)
        hoja_banco['M' + str(row)].value = n

    # for cell in hoja_banco[5]:
    #     cell.style = header

    # for file in deposito_file, extracto_file, emitidos_file, deposito_file:
    #     if file:
    #         os.remove(file)
    
    
    librobanco.save(banco)
    
    
    # #libro_banco.close()

    






if __name__ == '__main__':
    proceso()
































































# borro los xlsx
# for planilla in aceptados, deposito, emitidos:
#     os.system(f"rm {planilla}"+'x')

#  obtener los tipos de datos de la columna
# libro_banco_depositados.info()
# deposito.info()

#  dos fomras de sacar las diferencias. la segunda agrega una columna que indica donde esta.
# busca lo que esta en el  dataframe de adelante que este en el 2
# comparar = deposito2[~deposito2.apply(tuple, 1).isin(deposito.apply(tuple, 1))]
# aca compara y muestra las diferencias y lo q esta repetido segun lo q se ponga mejor para comparar ambos
# comparar = deposito.merge(deposito2, indicator=True, how='outer').loc[lambda v: v['_merge']!='both']
# depositos_faltantes = libro_banco_depositados.merge(deposito, indicator=True, how='outer').loc[lambda v: v['_merge']!='both']
